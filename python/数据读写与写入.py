import json
import random

from faker import Faker

"""

w write 读  r read 读
"""
# with open("txt.text","w") as f:
#     f.write("我得哩康迪\n")
#     f.write("我得发")
#     f.writelines(["123","123"])#一次多个写入

# with open("txt.text", "r") as f:
#     # print(f.readlines())#列表格式输出
#     lis = f.readlines()
#     for i in range(4):  # 输出四边
#         print(lis[1])  # 输出第一行
# print(f.readline())#每次读取一行
# print(f.readline())#每次读取一行

"""
faker

    zh_CN - Chinese (China Mainland)
    zh_TW - Chinese (China Taiwan)
    en_US - English (United States)


地理信息类
    city_suffix()：市，县
    country()：国家
    country_code()：国家编码
    district()：区
    geo_coordinate()：地理坐标
    latitude()：地理坐标(纬度)
    longitude()：地理坐标(经度)
    postcode()：邮编
    province()：省份
    address()：详细地址
    street_address()：街道地址
    street_name()：街道名
    street_suffix()：街、路
基础信息类
    ssn()：生成身份证号
    bs()：随机公司服务名
    company()：随机公司名（长）
    company_prefix()：随机公司名（短）
    company_suffix()：公司性质，如'信息有限公司'
    credit_card_expire()：随机信用卡到期日，如'03/30'
    credit_card_full()：生成完整信用卡信息
    credit_card_number()：信用卡号
    credit_card_provider()：信用卡类型
    credit_card_security_code()：信用卡安全码
    job()：随机职位
    first_name_female()：女性名
    first_name_male()：男性名
    name()：随机生成全名
    name_female()：男性全名
    name_male()：女性全名
    phone_number()：随机生成手机号
    phonenumber_prefix()：随机生成手机号段，如139
邮箱信息类
    ascii_company_email()：随机ASCII公司邮箱名
    ascii_email()：随机ASCII邮箱：
    company_email()：公司邮箱
    email()：普通邮箱
    safe_email()：安全邮箱
网络基础信息类
    domain_name()：生成域名
    domain_word()：域词(即，不包含后缀)
    ipv4()：随机IP4地址
    ipv6()：随机IP6地址
    mac_address()：随机MAC地址
    tld()：网址域名后缀(.com,.net.cn,等等，不包括.)
    uri()：随机URI地址
    uri_extension()：网址文件后缀
    uri_page()：网址文件（不包含后缀）
    uri_path()：网址文件路径（不包含文件名）
    url()：随机URL地址
    user_name()：随机用户名
    image_url()：随机URL地址
浏览器信息类
    chrome()：随机生成Chrome的浏览器user_agent信息
    firefox()：随机生成FireFox的浏览器user_agent信息
    internet_explorer()：随机生成IE的浏览器user_agent信息
    opera()：随机生成Opera的浏览器user_agent信息
    safari()：随机生成Safari的浏览器user_agent信息
    linux_platform_token()：随机Linux信息
    user_agent()：随机user_agent信息
数字信息
    numerify()：三位随机数字
    random_digit()：0~9随机数
    random_digit_not_null()：1~9的随机数
    random_int()：随机数字，默认0~9999，可以通过设置min,max来设置
    random_number()：随机数字，参数digits设置生成的数字位数
    pyfloat()：随机Float数字
    pyint()：随机Int数字（参考random_int()参数）
    pydecimal()：随机Decimal数字（参考pyfloat参数）
文本加密类
    pystr()：随机字符串
    random_element()：随机字母
    random_letter()：随机字母
    paragraph()：随机生成一个段落
    paragraphs()：随机生成多个段落
    sentence()：随机生成一句话
    sentences()：随机生成多句话，与段落类似
    text()：随机生成一篇文章
    word()：随机生成词语
    words()：随机生成多个词语，用法与段落，句子，类似
    binary()：随机生成二进制编码
    boolean()：True/False
    language_code()：随机生成两位语言编码
    locale()：随机生成语言/国际 信息
    md5()：随机生成MD5
    null_boolean()：NULL/True/False
    password()：随机生成密码,可选参数：length：密码长度；special_chars：是否能使用特殊字符；digits：是否包含数字；upper_case：是否包含大写字母；lower_case：是否包含小写字母
    sha1()：随机SHA1
    sha256()：随机SHA256
    uuid4()：随机UUID   
"""
# def MakFakeData():
import csv
import pandas as pd


# def fakeData():
#     data = [
#         ("姓名", "性别", "国家", "身份证号", "密码", "语文", "数学", "英语"),
#         ("王菲", "女", "中国", "123", "111a", 90, 80, 100),
#         ("王某飞", "男", "中国", "123", "111a", 60, 55, 50)
#     ]
#     xesLis = ['男', '女']
# #
#     for i in range(10):
#         r = random.randint(0, 1)
#         Name = fakerInfo.name()
#         sex = xesLis[r]
#         country = fakerInfo.country()
#         ssn = fakerInfo.ssn()
#         tup = (Name, sex, country, ssn, random.randint(0, 120), random.randint(0, 120), random.randint(0, 120))
#         data.append(tup)
#     with open('个人信息1.csv', 'w', newline='', encoding='utf-8')as f:
#         write = csv.writer(f)
#         for i in data:
#             write.writerow(i)


# def fakeDataWrite():
#     with open('个人信息1.csv', 'w', newline='', encoding='utf-8')as f:
#         fieldName = ['姓名', '语文', '数学', '英语']
#         writr = csv.DictWriter(f, fieldnames=fieldName)
#         writr.writeheader()
#         writr.writerow({
#             '姓名': fakerInfo.name(),
#             '语文': random.randint(0, 120),
#             '数学': random.randint(0, 120),
#             '英语': random.randint(0, 120)
#         })
#         writr.writerow({
#             '姓名': fakerInfo.name(),
#             '语文': random.randint(0, 120),
#             '数学': random.randint(0, 120),
#             '英语': random.randint(0, 120)
#         })


# def read():
#     with open("个人信息1.csv",'r',newline='', encoding='utf-8')as r:
#         reader=csv.reader(r)
#         print(reader)


def serializationAndDeserialization():
    str = '''
        [
            {
                "name":"小明",
                "en":10,
                "guojia":"Uk"
            },
            {
                "name": "xiaohua1",
                "en": 20,
                "guojia": "Uk"
            },
            {
                "name": "xiaohua1",
                "en": 30,
                "guojia": "UK"
            },
            {
                "name": "xiaohua1",
                "en": 40,
                "guojia": "UK"
            }
        ]
        '''
    # data = json.loads(str)
    # # print(data,type(str))
    # # datas = json.dumps(data)
    # # print(datas, type(datas))
    # print(data[0]['name'])
    # print(data[2]['guojia'])
    # print("---------------------------")
    # print(data[0].get('nam'))
    # print(data[0].get('nem', '张仨'))
    # print(data[0].get('naem', '张仨'))


import json

# def JsonData():
#     jsondata = [
#         {"姓名": "王菲", "年龄": 20, "数学": 10},
#         {"姓名": "小郭", "年龄": 20, "数学": 100},
#         {"姓名": "老康头", "年龄": 20, "数学": 1, "data": {
#             "data1": 1,
#             "data2": 2}
#          }
#
#     ]
#
#     with open("D:\Desktop\datajosn.json", "w", encoding="utf-8")as f:
#         f.write(json.dumps(jsondata, ensure_ascii=False, indent=2))
#         # 默认是ASCII编码形式insden=是用来添加缩进单位以空格为单位

import xlwt
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font

# def XlsxFileData():
#     wb = Workbook()
#     # she1 = wb.create_sheet('学元信息')  # 实例化
#     # she2 = wb.create_sheet('成绩信息')
#     # she1.append(["姓名", "年龄", "gender"])  # 添加数据
#     # she1.append(["王菲", "80", "女"])
#     # she1.append(["老康头", "120", "中性"])
#     # she1.append(["小郭", "18", "女"])
#     # she2.append(["语文", "数学", "英语"])
#     # she1["F10"] = "123"
#     # she1.merge_cells('H6:J14')  # 合并单元格
#     # she1.merge_cells('A{}:B{}'.format(5, 6))
#     # cell = she1.cell(row=8, column=4)
#     # cell.value = "我是针对你"
#     # cell.alignment = Alignment(horizontal="center", vertical="center")
#     # del wb["Sheet"]  # 删除
#     she2 = wb.create_sheet("成绩表")
#     she3 = wb.create_sheet("学院表")
#
#     she2.append(["语文", "数学", "英语"])
#     she3.append(["姓名", "学号", "班级"])
#
#     she3.append(["王菲", "003", "大二"])
#     she3.append(["王某", "002", "大三"])
#     she3.append(["菲某", "001", "大一"])
#     # 合并单元格
#     she2.merge_cells('A4:B5')
#     she2.merge_cells('C6:D7')
#     she2.merge_cells('A{}:B{}'.format(13, 17))
#     # 对单个单元格进行赋值
#     she3['B2'] = '007'
#     # 输出所有的工作簿
#     del wb["Sheet"]  # 删除
#     print(wb.sheetnames)  # 输出单元格
#     cell = she2.cell(row=13, column=1)
#
#     wb.save(r'D:\Desktop\nameinfo.xlsx')  # 输出路径
#     pass
#
#
# def XlsxLoadData():
#     wb = load_workbook(r'D:\Desktop\nameinfo.xlsx')
#     print(wb.sheetnames)
#     sh = wb['学院表']
#     print(sh.max_row, sh.max_column)
#     print(sh["A2"].value)
#     sh[1][2].value = "单元格修改"
#     wb.save(r'D:\Desktop\nameinfo.xlsx')
#
#
# def XlscxChangeStyle():
#     wb = load_workbook(r'D:\Desktop\nameinfo.xlsx')
#     print(wb.sheetnames)
#     sh = wb['学院表']
#     fontInfo = Font(name="微软雅黑", size=20, bold=True, color="F0F8FF")
#     sh['C1'].font = fontInfo
#     sh.row_dimensions[3].height = 30  # 行高
#     sh.column_dimensions['B'].width = 20  # 列宽
# def xlsFileData():
#     data = [
#         ("水星", 12, 12),
#         ("金星", 12, 12),
#         ("地球", 12, 12),
#         ("火星", 12, 12),
#         ("木星", 12, 12),
#         ("土星", 12, 12),
#         ("天王星", 12, 12),
#         ("海王星", 12, 12),
#     ]
#     book = xlwt.Workbook()  # 创建一个workbook对象
#     sh = book.add_sheet("太阳系八大行星")
#     col_name = ["行星名", "距离", "与地球质量比"]
#     for col, name in enumerate(col_name):
#         sh.write(0, col, name)
#         sh.col(col).width = 256 * 20
#
#     for i, line in enumerate(data):
#         for j, item in enumerate(line):
#             sh.write(i + 1, j, item)
#             pass
#
#     book.save(r"D:\Desktop\nameinfo.xls")
#
#
# def xlsLoadData():
#     book = xlrd.open_workbook(r"D:\Desktop\nameinfo.xls")
#     # print(book.sheet_by_names())#获取工作簿名
#     sheet = book.sheet_by_name("太阳系八大行星")  #
#     print(sheet.nrows)  # 输出有效行数
#     print(sheet.ncols)  # 输出有效列数
#     print(sheet.cell_value(3, 0))  # 从零开始
#     print(sheet.row(3))
#     print(sheet.cell(3, 0))


if __name__ == '__main__':
    fakerInfo = Faker(locale="zh_CN")
    # fakeData()
    # fakeDataWrite()
    # read()
    # serializationAndDeserialization()
    # JsonData()
    # XlsxFileData()
    # XlsxLoadData()
    # XlscxChangeStyle()
    # xlsFileData()
    # xlsLoadData()

all_products = []
fakeer = Faker(locale='zh_CN')
for i in range(10000):
    Name = fakeer.name()
    address = fakeer.address()
    ssn = fakeer.ssn()
    prefix = fakeer.company_prefix()
    number = fakeer.phone_number()
    number = fakeer.credit_card_number()

    all_products.append({  # 创建excel表
        "姓名": Name,
        "地址": address,
        "身份证号": ssn,
        "公司": prefix,
        "手机号": number,
        "信用卡号": number
    })
keys = all_products[0].keys()

with open('个人信息.txt', 'w', newline='', encoding='utf-8') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(all_products)

# 写入数据
# pd.DataFrame(all_products,columns=keys).to_csv('个人信息.csv', encoding='utf-8-sig')

"""
providers类名	作用
base	基础包，包含各种数字字母随机方法
address	地址相关
automotive	汽车行业
bank	银行
barcode	条码
color	颜色
company	公司（职称，口号）
credit_card	信用卡
currency	货币
date_time	日期时间
file	文件（扩展名，文件名，带路径文件名）
geo	地理
internet	互联网（邮箱，域名，hostname，图片url，ipv4、ipv6地址，mac地址）
isbn	书号
job	工作
lorem	文章句子
misc	杂项（md5，二进制串，密码，sha1，sha256，uuid）
person	人相关（人名（可分男女），姓，名，）
phone_number	电话号码（手机号，电话号码）
profile	制造JSON格式的数据
python	python相关（小数，列表，字典，集合，元组）
ssn	搞不懂
user_agent	用户代理（安卓设备渠道token，各类浏览器标识，操作系统token）
"""

# from faker import Faker
#
# fake=Faker() #默认生成美国英文数据,加入local参数指定生成数据是中文
# fake = Faker(locale='zh_CN')
# #
# # 1.个人信息相关数据
# print("个人信息类".center(20, "-"))
# # 东浩
# print(fake.name())
# # {'username': 'leihan', 'name': '武帅', 'sex': 'F', 'address': '吉林省淮安市双滦家街C座 210434', 'mail': 'lishao@hotmail.com', 'birthdate': '1988-11-12'}
# print(fake.simple_profile())
# # ajiang zI2QbHy02p
# print(fake.user_name(), fake.password(special_chars=False))
#
# # 2.地址相关数据
# print("地址类".center(20, "-"))
# # 海南省成市丰都深圳路p座 425541
# print(fake.address())
# # 深圳街X座
# print(fake.street_address())
# # 长沙路
# print(fake.street_name())
# # 兰州 贵阳市  (相差“市”)
# print(fake.city_name(), fake.city())
# # 陕西省
# print(fake.province())
#
# # 3. 公司信息相关数据
# # 公司名，比如星辰科技网络有限公司
# fake.company()
# # 公司名后缀(公司性质)，比如网络有限公司
# fake.company_suffix()
# # 公司名前缀，比如星辰科技
# fake.company_prefix()
#
# # 4.银行相关数据
# # 'BJZS4509618542924'
# fake.bban()
# # '北京'
# fake.bank_country()
# # 'GB04BPNH0448315286040'
# fake.iban()
#
# # 5.信用卡相关数据
# ## '360'
# fake.credit_card_security_code(card_type=None)
# # 'Diners Club / Carte Blanche\n刘 海\n30311852484679 10/19\nCVC: 388\n'
# fake.credit_card_full(card_type=None)
# # '30240280288941'
# fake.credit_card_number(card_type=None)
# # '11/26'
# fake.credit_card_expire(start="now", end="+10y", date_format="%m/%y")
# # 'Maestro'
# fake.credit_card_provider(card_type=None)
#
# # 6.条形码相关数据
# # # '0884331656275'
# fake.ean(length=13)
# # '52009350'
# fake.ean8()
# # '8766323543385'
# fake.ean13()
#
# # 7.文章相关数据
# print("文章类".center(20, "-"))
# print(fake.word())  # 当前
# print(fake.words(3))  # ['欢迎', '支持', '图片']
# print(fake.sentence(3))  # 精华有关一些.
# # 成功的花儿，人们只惊艳它现实的明艳，然而当初它芽，浸透了奋斗的潜水，洒遍了牺牲的血雨.
# print(fake.paragraph())
#
# # 8.日期相关数据
# fake.date(pattern="%Y-%m-%d", end_datetime=None)
# # 随机年份
# fake.year()
# # 随机星期数
# fake.day_of_week()
# # 随机时间
# fake.time(pattern="%H:%M:%S", end_datetime=None)
#
# # 9.网络相关数据
# # 企业邮箱
# fake.company_email()
# # 邮箱
# fake.email()
